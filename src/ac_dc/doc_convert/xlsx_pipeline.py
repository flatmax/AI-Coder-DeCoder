"""Colour-aware xlsx → markdown pipeline using openpyxl. Extracted from the original monolithic `doc_convert.py` during the package split."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from .constants import (
    _COLOUR_CLUSTER_DISTANCE,
    _FALLBACK_MARKERS,
    _IGNORE_NEAR_BLACK_THRESHOLD,
    _IGNORE_NEAR_WHITE_THRESHOLD,
    _NAMED_COLOURS,
    _NAMED_COLOUR_DISTANCE,
)
from .provenance import build_provenance_header, hash_file

logger = logging.getLogger(__name__)


class XlsxPipeline:
    def __init__(self, fail, skip, markitdown_fallback) -> None:
        self._fail = fail
        self._skip = skip
        self._markitdown_fallback = markitdown_fallback

    def convert(
        self,
        root: Path,
        source_abs: Path,
        rel_path: str,
    ) -> dict[str, Any]:
        """Convert an .xlsx file, preserving cell background colours.

        Two-pass approach:

        1. Pass 1 — walk every cell in every sheet, collecting
           text values (normalised) and raw hex fill colours. The
           set of unique non-ignorable fills is built during this
           pass.
        2. Colour mapping — well-known hues (red, green, yellow,
           blue, purple, etc.) get named emoji markers. Remaining
           colours are clustered by Euclidean RGB distance and
           assigned fallback markers per cluster.
        3. Pass 2 — emit markdown tables sheet by sheet, cells
           prefixed with their colour marker.

        Empty columns and fully-empty rows are stripped. A
        legend mapping markers to colour names appears at the end.

        Falls back to markitdown on any openpyxl failure
        (ImportError, corrupt file, unexpected structure) — the
        user still gets SOMETHING rather than an error result.
        """
        # Lazy import — openpyxl is optional in stripped-down
        # releases. ImportError is expected in that case and
        # means "use markitdown instead", not "fail the
        # conversion".
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.debug(
                "openpyxl not installed; falling back to "
                "markitdown for %s",
                rel_path,
            )
            return self._markitdown_fallback(
                root, source_abs, rel_path
            )

        output_abs = source_abs.with_suffix(".md")
        try:
            output_rel = output_abs.relative_to(root)
        except ValueError:
            return self._fail(
                rel_path,
                "Output path escapes repository root",
            )

        # Hash source for provenance — matches the markitdown
        # path so status classification works uniformly.
        try:
            source_hash = hash_file(source_abs)
        except OSError as exc:
            return self._fail(
                rel_path, f"Source hash failed: {exc}"
            )

        # Open workbook in read-only mode for performance on
        # large files. Also pass data_only=False so we see
        # formulas as formulas (not cached values) — mostly
        # defensive; formula cells rarely have fills anyway.
        try:
            workbook = load_workbook(
                filename=str(source_abs),
                read_only=False,  # need .fill on cells
                data_only=True,
            )
        except Exception as exc:
            # Corrupt xlsx, password-protected, etc. — fall back
            # to markitdown rather than erroring. markitdown may
            # still extract something useful.
            logger.debug(
                "openpyxl failed to open %s: %s; "
                "falling back to markitdown",
                rel_path, exc,
            )
            return self._markitdown_fallback(
                root, source_abs, rel_path
            )

        # Pass 1: collect cells and unique fills across all sheets.
        try:
            sheets_data, unique_fills = self._xlsx_pass1_collect(
                workbook
            )
        except Exception as exc:
            logger.debug(
                "openpyxl pass-1 failed for %s: %s; "
                "falling back to markitdown",
                rel_path, exc,
            )
            workbook.close()
            return self._markitdown_fallback(
                root, source_abs, rel_path
            )
        finally:
            workbook.close()

        # Build colour → marker map.
        colour_map = self._xlsx_build_colour_map(unique_fills)

        # Pass 2: emit markdown per sheet.
        body_parts: list[str] = []
        for sheet_name, rows in sheets_data:
            sheet_md = self._xlsx_render_sheet(
                sheet_name, rows, colour_map
            )
            if sheet_md:
                body_parts.append(sheet_md)

        if not body_parts:
            # Workbook had no data in any sheet. Still produce
            # an output file so the scan classifies it as
            # `current`, but with an informative placeholder.
            body_parts.append("(empty spreadsheet)")

        # Append legend mapping each used marker to its colour name.
        legend = self._xlsx_render_legend(colour_map)
        if legend:
            body_parts.append(legend)

        markdown_text = "\n\n".join(body_parts) + "\n"

        # Write output with provenance header. xlsx path never
        # produces embedded images, so we skip the data-URI
        # extraction pipeline entirely.
        provenance_line = build_provenance_header(
            source_name=source_abs.name,
            source_hash=source_hash,
            images=(),
        )
        final_content = (
            provenance_line + "\n\n" + markdown_text.lstrip("\n")
        )

        try:
            output_abs.parent.mkdir(parents=True, exist_ok=True)
            output_abs.write_text(final_content, encoding="utf-8")
        except OSError as exc:
            return self._fail(
                rel_path,
                f"Failed to write output: {exc}",
            )

        return {
            "path": rel_path,
            "status": "ok",
            "output_path": str(output_rel).replace("\\", "/"),
            "images": [],
        }

    def _xlsx_pass1_collect(
        self,
        workbook: Any,
    ) -> tuple[
        list[tuple[str, list[list[tuple[str, str | None]]]]],
        set[str],
    ]:
        """First pass over an xlsx workbook.

        Returns a list of `(sheet_name, rows)` pairs and the set
        of unique non-ignorable hex fill colours across the
        workbook. Each row is a list of `(value, hex_fill)`
        tuples where `hex_fill` is None for ignorable fills.

        Cell values are normalised — None becomes empty string,
        "nan"/"none" (case-insensitive) become empty string, and
        all non-string values are stringified. Whitespace is
        preserved except that leading/trailing is trimmed.

        Exceptions from openpyxl propagate; the caller wraps
        with a fallback-to-markitdown path.
        """
        sheets_data: list[
            tuple[str, list[list[tuple[str, str | None]]]]
        ] = []
        unique_fills: set[str] = set()

        for sheet in workbook.worksheets:
            rows: list[list[tuple[str, str | None]]] = []
            # iter_rows with values_only=False gives us Cell
            # objects (needed for .fill). We walk the used range
            # only — iter_rows defaults to the sheet's dimension.
            for row in sheet.iter_rows():
                row_cells: list[tuple[str, str | None]] = []
                for cell in row:
                    value = self._normalise_cell_value(cell.value)
                    fill_hex = self._extract_cell_fill(cell)
                    if fill_hex is not None:
                        unique_fills.add(fill_hex)
                    row_cells.append((value, fill_hex))
                rows.append(row_cells)
            sheets_data.append((sheet.title, rows))

        return sheets_data, unique_fills

    @staticmethod
    def _normalise_cell_value(value: Any) -> str:
        """Normalise a raw cell value for markdown emission.

        - None → empty string
        - Pandas/numpy artifacts "nan" / "none" (case-insensitive)
          → empty string. These crop up when a spreadsheet was
          generated from a DataFrame with missing values.
        - Everything else is stringified and stripped of
          leading/trailing whitespace.

        Pipe characters are escaped as `\\|` since they would
        otherwise break the markdown table row.
        """
        if value is None:
            return ""
        text = str(value).strip()
        if text.lower() in ("nan", "none"):
            return ""
        # Escape pipes so they don't break table rows. Literal
        # backslashes in cell values are rare; we don't escape
        # those.
        return text.replace("|", r"\|")

    @staticmethod
    def _extract_cell_fill(cell: Any) -> str | None:
        """Return the cell's fill as a hex string, or None.

        Ignorable fills (near-white, near-black, no fill at all)
        return None so they don't produce emoji markers. The
        hex string is lowercase without a leading hash.

        openpyxl's fill model is verbose — cells with no explicit
        fill still have a PatternFill with fgColor set to a
        default theme colour. We filter those by checking the
        patternType and the raw RGB.
        """
        try:
            fill = cell.fill
            if fill is None:
                return None
            # Only solid fills carry meaningful colour info.
            # patternType "none" means no explicit fill.
            pattern_type = getattr(fill, "patternType", None)
            if pattern_type not in ("solid", "lightGrid", "darkGrid"):
                return None
            fg = fill.fgColor
            if fg is None:
                return None
            # fgColor.rgb is an 8-char hex string (AARRGGBB)
            # when set. Theme colours return None here; we can't
            # resolve them without the workbook's theme table,
            # which isn't worth the complexity for a diagnostic
            # marker.
            raw = getattr(fg, "rgb", None)
            if not raw or not isinstance(raw, str):
                return None
            # Some versions of openpyxl return the rgb as a
            # Value wrapper — coerce defensively.
            raw = str(raw).strip().lower()
            # Strip alpha channel if present.
            if len(raw) == 8:
                raw = raw[2:]
            if len(raw) != 6:
                return None
            # Filter ignorable colours.
            try:
                r = int(raw[0:2], 16)
                g = int(raw[2:4], 16)
                b = int(raw[4:6], 16)
            except ValueError:
                return None
            # Near-white → ignore.
            if (
                (255 - r) < _IGNORE_NEAR_WHITE_THRESHOLD
                and (255 - g) < _IGNORE_NEAR_WHITE_THRESHOLD
                and (255 - b) < _IGNORE_NEAR_WHITE_THRESHOLD
            ):
                return None
            # Near-black → ignore.
            if (
                r < _IGNORE_NEAR_BLACK_THRESHOLD
                and g < _IGNORE_NEAR_BLACK_THRESHOLD
                and b < _IGNORE_NEAR_BLACK_THRESHOLD
            ):
                return None
            return raw
        except Exception:
            # Anything unexpected — treat as no fill. Defensive
            # against openpyxl API drift.
            return None

    @staticmethod
    def _hex_to_rgb(hex_str: str) -> tuple[int, int, int]:
        """Convert a 6-char hex string to an (r, g, b) tuple.

        Input is assumed valid (produced by `_extract_cell_fill`
        which already validates). No error handling on the
        integer parses — if they fail, the caller has a bug.
        """
        return (
            int(hex_str[0:2], 16),
            int(hex_str[2:4], 16),
            int(hex_str[4:6], 16),
        )

    @staticmethod
    def _colour_distance(
        a: tuple[int, int, int],
        b: tuple[int, int, int],
    ) -> float:
        """Euclidean RGB distance.

        Perceptually naive (doesn't weight green higher like
        proper colour-diff metrics) but fine for the "are these
        two reds the same colour?" question the clustering needs.
        """
        dr = a[0] - b[0]
        dg = a[1] - b[1]
        db = a[2] - b[2]
        return (dr * dr + dg * dg + db * db) ** 0.5

    def _xlsx_build_colour_map(
        self,
        unique_fills: set[str],
    ) -> dict[str, tuple[str, str]]:
        """Assign an emoji marker and colour name to each fill.

        Returns a dict mapping hex colour → `(marker, name)`
        tuple. The name is either a named-colour label (red,
        green, etc.) or a synthesised label for fallback
        clusters ("cluster-1", "cluster-2", …).

        Algorithm:

        1. For each unique fill, find the closest named colour
           within `_NAMED_COLOUR_DISTANCE`. If found, assign the
           named marker.
        2. For remaining fills (no named match), cluster by
           proximity — fills within `_COLOUR_CLUSTER_DISTANCE` of
           an existing cluster join it; otherwise start a new
           cluster. Assign fallback markers in order.

        Named colours can be shared by multiple fills (all
        "reddish" cells get 🔴 regardless of exact shade).
        Fallback clusters each get their own marker.
        """
        result: dict[str, tuple[str, str]] = {}

        # Sort fills for deterministic assignment — two runs on
        # the same workbook produce the same markers. Without
        # this, set iteration order would vary.
        sorted_fills = sorted(unique_fills)

        unmatched: list[tuple[str, tuple[int, int, int]]] = []

        # Step 1 — assign named colours where close enough.
        for hex_fill in sorted_fills:
            rgb = self._hex_to_rgb(hex_fill)
            best_name: str | None = None
            best_marker: str | None = None
            best_dist = _NAMED_COLOUR_DISTANCE
            for name, named_rgb, marker in _NAMED_COLOURS:
                dist = self._colour_distance(rgb, named_rgb)
                if dist < best_dist:
                    best_dist = dist
                    best_name = name
                    best_marker = marker
            if best_name is not None and best_marker is not None:
                result[hex_fill] = (best_marker, best_name)
            else:
                unmatched.append((hex_fill, rgb))

        # Step 2 — cluster remaining fills. Each cluster holds a
        # representative RGB and the set of hex strings assigned
        # to it.
        clusters: list[tuple[tuple[int, int, int], list[str]]] = []
        for hex_fill, rgb in unmatched:
            joined = False
            for i, (cluster_rgb, cluster_hexes) in enumerate(clusters):
                if self._colour_distance(rgb, cluster_rgb) < _COLOUR_CLUSTER_DISTANCE:
                    cluster_hexes.append(hex_fill)
                    joined = True
                    break
            if not joined:
                clusters.append((rgb, [hex_fill]))

        # Assign fallback markers. More clusters than we have
        # markers — cycle through and append an index to the
        # name so entries remain distinguishable in the legend.
        for i, (_rgb, cluster_hexes) in enumerate(clusters):
            marker = _FALLBACK_MARKERS[i % len(_FALLBACK_MARKERS)]
            name = f"cluster-{i + 1}"
            for hex_fill in cluster_hexes:
                result[hex_fill] = (marker, name)

        return result

    def _xlsx_render_sheet(
        self,
        sheet_name: str,
        rows: list[list[tuple[str, str | None]]],
        colour_map: dict[str, tuple[str, str]],
    ) -> str:
        """Render one sheet as a markdown section.

        - Empty columns (all values empty across every row) are
          dropped.
        - Fully-empty rows are dropped.
        - First non-empty row becomes the table header. If every
          cell in that row is a string with meaningful content,
          it's used as-is; otherwise synthetic headers
          (`col1`, `col2`, …) are generated.
        - Coloured cells get their marker prepended with a
          space separator.

        Returns an empty string when the sheet has no data after
        stripping — the caller skips those sheets.
        """
        # Drop fully-empty rows.
        non_empty_rows = [
            row for row in rows
            if any(value for value, _ in row)
        ]
        if not non_empty_rows:
            return ""

        # Find the widest row — column count is the max width
        # across all non-empty rows. Shorter rows are padded
        # with empty cells during render.
        max_width = max(len(row) for row in non_empty_rows)
        if max_width == 0:
            return ""

        # Drop fully-empty columns. A column is empty if every
        # non-empty row has an empty value at that position.
        keep_columns: list[int] = []
        for col_idx in range(max_width):
            for row in non_empty_rows:
                if col_idx < len(row) and row[col_idx][0]:
                    keep_columns.append(col_idx)
                    break

        if not keep_columns:
            return ""

        # Header row — use the first non-empty row's values if
        # they look like headers (all non-empty strings, no
        # colour markers). Otherwise synthesise column names.
        first_row = non_empty_rows[0]
        use_first_as_header = all(
            col_idx < len(first_row) and first_row[col_idx][0]
            for col_idx in keep_columns
        )

        header_cells: list[str]
        data_rows: list[list[tuple[str, str | None]]]
        if use_first_as_header:
            header_cells = [
                first_row[col_idx][0] for col_idx in keep_columns
            ]
            data_rows = non_empty_rows[1:]
        else:
            header_cells = [
                f"col{i + 1}" for i in range(len(keep_columns))
            ]
            data_rows = non_empty_rows

        # Build markdown.
        lines: list[str] = [f"## {sheet_name}", ""]
        lines.append(
            "| " + " | ".join(header_cells) + " |"
        )
        lines.append(
            "|" + "|".join("---" for _ in keep_columns) + "|"
        )
        for row in data_rows:
            rendered_cells: list[str] = []
            for col_idx in keep_columns:
                if col_idx < len(row):
                    value, fill_hex = row[col_idx]
                else:
                    value, fill_hex = "", None
                if fill_hex is not None and fill_hex in colour_map:
                    marker = colour_map[fill_hex][0]
                    rendered_cells.append(
                        f"{marker} {value}" if value else marker
                    )
                else:
                    rendered_cells.append(value)
            lines.append(
                "| " + " | ".join(rendered_cells) + " |"
            )

        return "\n".join(lines)

    @staticmethod
    def _xlsx_render_legend(
        colour_map: dict[str, tuple[str, str]],
    ) -> str:
        """Render the colour-marker legend at the end of the output.

        Lists each unique (marker, name) pair exactly once. The
        colour map may contain multiple hex values mapped to the
        same named colour (all reddish fills → 🔴 red); the
        legend shows the named entry once rather than repeating
        per-hex.

        Returns an empty string when no markers were used.
        """
        if not colour_map:
            return ""
        # Collect unique (marker, name) pairs.
        seen: set[tuple[str, str]] = set()
        ordered_entries: list[tuple[str, str]] = []
        for marker, name in colour_map.values():
            key = (marker, name)
            if key in seen:
                continue
            seen.add(key)
            ordered_entries.append(key)
        if not ordered_entries:
            return ""
        lines = ["## Legend", ""]
        for marker, name in ordered_entries:
            lines.append(f"- {marker} {name}")
        return "\n".join(lines)